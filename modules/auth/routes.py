import io
import os
import re
from flask import Blueprint, render_template, redirect, url_for, request, jsonify, send_file, session
from flask_login import login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash

from database import db
from database.models import User, NotificationPreference
from extensions import limiter
from .decorators import require_role
from utils.validators import validate_password_strength
from utils.audit import log_action
from utils.encryption import compute_search_hash

# Hash dummy precomputado para neutralizar el timing oracle del login.
# Why: cuando un usuario no existe, saltarse la verificación del password
# devuelve la respuesta más rápido que cuando sí existe (porque el hash check
# tarda ~100 ms). Comparando latencias se puede enumerar usuarios válidos.
# How to apply: en login(), ejecutar SIEMPRE check_password_hash, contra el
# hash real si el usuario existe o contra _DUMMY_PASSWORD_HASH si no.
_DUMMY_PASSWORD_HASH = generate_password_hash('___never_matches___')

_EMAIL_RE = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')
_COLOR_RE = re.compile(r'^#[0-9A-Fa-f]{6}$')
_ALLOWED_LANGS = {'es', 'en'}
_ALLOWED_THEMES = {'light', 'dark'}

auth_bp = Blueprint('auth', __name__, url_prefix='/auth')

VALID_ROLES = ('superadmin', 'admin', 'operador', 'solo_lectura')
_USERNAME_RE = re.compile(r'^[a-zA-Z0-9._\-]{2,80}$')


# ── Login / Logout ────────────────────────────────────────────────────────────

@auth_bp.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute; 20 per hour", methods=['POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()

        # Verificación en tiempo constante: ejecuta SIEMPRE check_password_hash
        # contra un hash real (del usuario) o dummy (si no existe) para que el
        # tiempo de respuesta no permita enumerar usuarios válidos.
        password_hash = user.password_hash if user else _DUMMY_PASSWORD_HASH
        password_ok = check_password_hash(password_hash, password)

        if user is None:
            log_action('login_failed', f'user:{username}')
            error = 'Usuario o contraseña incorrectos'
        elif user.locked_until and user.locked_until > datetime.utcnow():
            mins = max(1, int((user.locked_until - datetime.utcnow()).total_seconds() / 60) + 1)
            log_action('login_blocked', f'user:{username}')
            error = f'Cuenta bloqueada por demasiados intentos fallidos. Inténtalo de nuevo en {mins} minuto{"s" if mins != 1 else ""}.'
        elif user.is_active and password_ok:
            user.failed_attempts = 0
            user.locked_until = None
            user.last_login = datetime.utcnow()
            db.session.commit()
            login_user(user, remember=False)
            session.permanent = True
            session['_last_active'] = datetime.utcnow().isoformat()
            log_action('login_success', f'user:{username}')
            next_url = request.args.get('next', '')
            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            return redirect(url_for('index'))
        else:
            user.failed_attempts = (user.failed_attempts or 0) + 1
            if user.failed_attempts >= 5:
                user.locked_until = datetime.utcnow() + timedelta(minutes=15)
            db.session.commit()
            log_action('login_failed', f'user:{username}')
            error = 'Usuario o contraseña incorrectos'

    return render_template('auth/login.html', error=error)


@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))


# ── Página gestión usuarios ───────────────────────────────────────────────────

@auth_bp.route('/usuarios')
@require_role('superadmin', 'admin')
def usuarios_page():
    users = User.query.order_by(User.created_at).all()
    return render_template('auth/usuarios.html', users=users, now_utc=datetime.utcnow())


# ── API CRUD usuarios ─────────────────────────────────────────────────────────
# - Crear/Eliminar: sólo superadmin
# - Listar/Editar (incluido cambio de username): admin + superadmin
#   Admin tiene restricciones: no puede modificar superadmins ni cambiar roles.

@auth_bp.route('/api/usuarios', methods=['GET'])
@require_role('superadmin', 'admin')
def list_users():
    users = User.query.order_by(User.created_at).all()
    return jsonify([u.to_dict() for u in users])


@auth_bp.route('/api/usuarios', methods=['POST'])
@require_role('superadmin')
def create_user():
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    role = data.get('role', 'solo_lectura')
    email = (data.get('email') or '').strip().lower() or None
    full_name = (data.get('full_name') or '').strip() or None

    if not username or not password:
        return jsonify({'error': 'Se requieren username y password'}), 400
    if role not in VALID_ROLES:
        return jsonify({'error': 'Rol no válido'}), 400
    if not _USERNAME_RE.match(username):
        return jsonify({'error': 'Nombre de usuario no válido (2-80 caracteres: letras, números, ., _, -)'}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'El nombre de usuario ya existe'}), 409

    if email:
        if not _EMAIL_RE.match(email):
            return jsonify({'error': 'Correo electrónico no válido'}), 400
        if User.query.filter_by(email_hash=compute_search_hash(email)).first():
            return jsonify({'error': 'Ese correo ya está en uso por otro usuario'}), 409

    if full_name and len(full_name) > 120:
        return jsonify({'error': 'El nombre no puede superar 120 caracteres'}), 400

    try:
        validate_password_strength(password)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    user = User(
        username=username,
        role=role,
        is_active=True,
        email=email,
        email_hash=compute_search_hash(email),
        full_name=full_name,
    )
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    log_action('user_created', f'user:{username}',
               f'role={role} email={email or "-"}')

    # Notificación de bienvenida (no bloquea: encola en EmailQueue).
    # Import diferido a propósito: notifications.dispatch importa el blueprint
    # de notifications, que a su vez importa modelos que dependen de auth.
    # Hacerlo top-level crea un ciclo en el arranque de la app.
    try:
        from modules.notifications.dispatch import notify_welcome
        notify_welcome(user)
    except Exception:                              # noqa: BLE001
        pass

    return jsonify(user.to_dict()), 201


@auth_bp.route('/api/usuarios/<int:user_id>', methods=['PUT'])
@require_role('superadmin', 'admin')
def update_user(user_id):
    user = db.session.get(User, user_id)
    if user is None:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    data = request.get_json(silent=True) or {}
    is_admin_only = (current_user.role == 'admin')

    # Admin no puede modificar cuentas superadmin
    if is_admin_only and user.role == 'superadmin' and user.id != current_user.id:
        return jsonify({'error': 'No tiene permisos para modificar superadmins'}), 403

    # USERNAME — admin y superadmin pueden cambiarlo
    if 'username' in data:
        new_username = (data['username'] or '').strip()
        if new_username != user.username:
            if not _USERNAME_RE.match(new_username):
                return jsonify({'error': 'Nombre de usuario no válido (2-80 caracteres: letras, números, ., _, -)'}), 400
            if User.query.filter(User.username == new_username, User.id != user.id).first():
                return jsonify({'error': 'Ese nombre de usuario ya existe'}), 409
            user.username = new_username

    # ROLE — sólo superadmin
    if 'role' in data:
        if is_admin_only:
            return jsonify({'error': 'Sólo superadmin puede cambiar roles'}), 403
        if data['role'] not in VALID_ROLES:
            return jsonify({'error': 'Rol no válido'}), 400
        if user.id == current_user.id and data['role'] != 'superadmin':
            return jsonify({'error': 'No puede cambiar su propio rol'}), 400
        user.role = data['role']

    # EMAIL — admin y superadmin pueden modificarlo
    if 'email' in data:
        new_email = (data['email'] or '').strip().lower() or None
        if new_email and not _EMAIL_RE.match(new_email):
            return jsonify({'error': 'Correo electrónico no válido'}), 400
        new_hash = compute_search_hash(new_email)
        if new_email and User.query.filter(
            User.email_hash == new_hash, User.id != user.id
        ).first():
            return jsonify({'error': 'Ese correo ya está en uso por otro usuario'}), 409
        user.email = new_email
        user.email_hash = new_hash

    # FULL_NAME — opcional, máx 120
    if 'full_name' in data:
        full_name = (data['full_name'] or '').strip() or None
        if full_name and len(full_name) > 120:
            return jsonify({'error': 'El nombre no puede superar 120 caracteres'}), 400
        user.full_name = full_name

    if data.get('password'):
        try:
            validate_password_strength(data['password'])
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        user.set_password(data['password'])

    if 'is_active' in data:
        if user.id == current_user.id:
            return jsonify({'error': 'No puede desactivarse a sí mismo'}), 400
        user.is_active = bool(data['is_active'])

    db.session.commit()
    log_action('user_updated', f'user_id:{user_id}', f'username={user.username} fields={list(data.keys())}')
    return jsonify(user.to_dict())


@auth_bp.route('/api/usuarios/<int:user_id>', methods=['DELETE'])
@require_role('superadmin')
def delete_user(user_id):
    user = db.session.get(User, user_id)
    if user is None:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    if user.id == current_user.id:
        return jsonify({'error': 'No puede eliminar su propia cuenta'}), 400
    username_deleted = user.username
    db.session.delete(user)
    db.session.commit()
    log_action('user_deleted', f'user_id:{user_id}', f'username={username_deleted}')
    return jsonify({'message': f'Usuario {user_id} eliminado'})


@auth_bp.route('/api/usuarios/<int:user_id>/unlock', methods=['POST'])
@require_role('superadmin', 'admin')
def unlock_user(user_id):
    """Quita el bloqueo de cuenta y resetea el contador de intentos fallidos.
    Admin no puede desbloquear cuentas superadmin (mismo criterio que update)."""
    user = db.session.get(User, user_id)
    if user is None:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    if current_user.role == 'admin' and user.role == 'superadmin' and user.id != current_user.id:
        return jsonify({'error': 'No tiene permisos para desbloquear superadmins'}), 403

    user.locked_until = None
    user.failed_attempts = 0
    db.session.commit()
    log_action('user_unlocked', f'user_id:{user_id}', f'username={user.username}')
    return jsonify(user.to_dict())


# ── API Perfil (usuario autenticado) ──────────────────────────────────────────

# ── Página y API de registros (audit log) ────────────────────────────────────

@auth_bp.route('/registros')
@require_role('superadmin', 'admin')
def registros_page():
    return render_template('auth/registros.html')


def _load_filtered_audit_entries(args):
    """Lee audit.log + rotaciones, aplica filtros y devuelve la lista
    filtrada (más reciente primero) + sets de actions/users/machines.

    Centraliza la lógica para ser reutilizada por list_audit_logs y export.
    """
    from utils.audit import AUDIT_LOG_PATH

    user_q = (args.get('user', '') or '').strip().lower()
    action_q = (args.get('action', '') or '').strip().lower()
    machine_q = (args.get('machine', '') or '').strip().lower()
    text_q = (args.get('q', '') or '').strip().lower()
    date_from = (args.get('from', '') or '').strip()
    date_to = (args.get('to', '') or '').strip()

    base = AUDIT_LOG_PATH
    paths = [base] if os.path.exists(base) else []
    for i in range(1, 6):
        p = f'{base}.{i}'
        if os.path.exists(p):
            paths.append(p)

    line_re = re.compile(
        r'^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}(?:,\d+)?) '
        r'user=(?P<user>\S+?)(?:\s+uid=(?P<uid>\d+))? ip=(?P<ip>\S+) '
        r'(?:machine=(?P<machine>\S+) )?'
        r'action=(?P<action>\S+) resource=(?P<resource>\S+)'
        r'(?: detail="(?P<detail>[^"]*)")?'
    )

    entries = []
    actions_set, users_set, machines_set = set(), set(), set()
    for path in paths:
        try:
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                for line in f:
                    m = line_re.match(line.strip())
                    if not m:
                        continue
                    machine = m.group('machine') or '-'
                    machine_disp = machine.replace('_', ' ') if machine != '-' else '-'
                    e = {
                        'timestamp': m.group('ts'),
                        'user':      m.group('user'),
                        'uid':       m.group('uid'),
                        'ip':        m.group('ip'),
                        'machine':   machine_disp,
                        'action':    m.group('action'),
                        'resource':  m.group('resource'),
                        'detail':    m.group('detail') or '',
                    }
                    actions_set.add(e['action'])
                    users_set.add(e['user'])
                    if machine_disp and machine_disp != '-':
                        machines_set.add(machine_disp)
                    entries.append(e)
        except (IOError, OSError):
            continue

    # Resolver username actual desde la BD usando el uid guardado en el log.
    # Si el usuario cambió su nombre, se muestra el nombre actual.
    try:
        from database.models import User as _User
        uid_set = {int(e['uid']) for e in entries if e.get('uid')}
        if uid_set:
            uid_map = {u.id: u.username for u in _User.query.filter(_User.id.in_(uid_set)).all()}
            for e in entries:
                if e.get('uid') and int(e['uid']) in uid_map:
                    e['user'] = uid_map[int(e['uid'])]
        # Reconstruir users_set con los nombres actuales tras la resolución
        users_set = {e['user'] for e in entries}
    except Exception:
        pass

    def _keep(e):
        if user_q and user_q not in e['user'].lower():
            return False
        if action_q and action_q != e['action'].lower():
            return False
        if machine_q and machine_q not in (e['machine'] or '').lower():
            return False
        if text_q:
            hay = ' '.join([e['user'], e['action'], e['resource'],
                            e['detail'], e['machine']]).lower()
            if text_q not in hay:
                return False
        ts_date = e['timestamp'][:10]
        if date_from and ts_date < date_from:
            return False
        if date_to and ts_date > date_to:
            return False
        return True

    filtered = [e for e in entries if _keep(e)]
    filtered.reverse()  # más reciente primero
    return filtered, actions_set, users_set, machines_set, base


@auth_bp.route('/api/registros', methods=['GET'])
@require_role('superadmin', 'admin')
def list_audit_logs():
    """Devuelve líneas del audit.log (más reciente primero), con filtros y paginación."""
    import logging
    log = logging.getLogger(__name__)
    try:
        try:
            page = max(int(request.args.get('page', 1)), 1)
            per_page = min(max(int(request.args.get('per_page', 50)), 10), 500)
        except (TypeError, ValueError):
            page, per_page = 1, 50

        filtered, actions_set, users_set, machines_set, _base = _load_filtered_audit_entries(request.args)

        total = len(filtered)
        start = (page - 1) * per_page
        end = start + per_page

        # No exponer la ruta del log al frontend (defensa en profundidad).
        return jsonify({
            'entries':  filtered[start:end],
            'total':    total,
            'page':     page,
            'per_page': per_page,
            'pages':    (total + per_page - 1) // per_page if per_page else 1,
            'actions':  sorted(actions_set),
            'users':    sorted(users_set),
            'machines': sorted(machines_set),
        })
    except Exception as e:
        log.exception("Error listando registros de auditoría")
        return jsonify({'error': f'{type(e).__name__}: {e}'}), 500


@auth_bp.route('/api/registros/export', methods=['GET'])
@require_role('superadmin', 'admin')
def export_audit_logs():
    """Exporta los registros filtrados en CSV / XLSX / PDF."""
    fmt = (request.args.get('format', 'csv') or '').lower()
    if fmt not in ('csv', 'xlsx', 'pdf'):
        return jsonify({'error': "Formato no válido. Use csv, xlsx o pdf."}), 400

    filtered, _, _, _, _ = _load_filtered_audit_entries(request.args)

    headers = ['Fecha y Hora', 'Usuario', 'Máquina', 'Acción', 'Recurso', 'IP origen', 'Detalle']
    rows = [
        [e['timestamp'].split(',')[0], e['user'], e['machine'],
         e['action'], e['resource'], e['ip'], e['detail']]
        for e in filtered
    ]

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname_base = f'registros_{ts}'

    log_action('audit_export', f'fmt={fmt}', f'count={len(rows)}')

    if fmt == 'csv':
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        writer.writerows(rows)
        # UTF-8 con BOM para que Excel lo abra bien
        data = ('﻿' + buf.getvalue()).encode('utf-8')
        return send_file(
            io.BytesIO(data),
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name=f'{fname_base}.csv',
        )

    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'Registros'
        ws.append(headers)
        # Estilo de cabecera
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='0A67DC')
        for c in ws[1]:
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='left', vertical='center')
        for row in rows:
            ws.append(row)
        # Anchos aproximados
        widths = [20, 16, 18, 22, 28, 14, 60]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = 'A2'
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{fname_base}.xlsx',
        )

    # PDF
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []
    title = Paragraph(f"Registros de la Aplicación — {len(rows)} entradas", styles['Heading2'])
    elements.append(title)
    elements.append(Spacer(1, 8))

    # Truncar campos muy largos para que quepan; convertir a Paragraph para wrap
    body_style = styles['BodyText']
    body_style.fontSize = 7
    body_style.leading = 9
    table_data = [headers] + [
        [Paragraph(str(c).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'), body_style) for c in r]
        for r in rows
    ]
    col_widths = [78, 60, 70, 80, 110, 56, 240]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0A67DC')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fb')]),
    ]))
    elements.append(table)
    doc.build(elements)
    out.seek(0)
    return send_file(
        out,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'{fname_base}.pdf',
    )


@auth_bp.route('/api/profile', methods=['GET'])
@login_required
def get_profile():
    return jsonify(current_user.to_dict())


@auth_bp.route('/api/profile/update', methods=['POST'])
@login_required
def update_profile():
    data = request.get_json(silent=True) or {}
    changed = []

    if 'email' in data:
        email = (data['email'] or '').strip().lower()
        if email and not _EMAIL_RE.match(email):
            return jsonify({'error': 'Correo electrónico no válido'}), 400
        prof_hash = compute_search_hash(email) if email else None
        if email and User.query.filter(
            User.email_hash == prof_hash, User.id != current_user.id
        ).first():
            return jsonify({'error': 'El correo ya está en uso por otro usuario'}), 409
        current_user.email = email or None
        current_user.email_hash = prof_hash
        changed.append('email')

    if 'full_name' in data:
        full_name = (data['full_name'] or '').strip()
        if len(full_name) > 120:
            return jsonify({'error': 'El nombre no puede superar 120 caracteres'}), 400
        current_user.full_name = full_name or None
        changed.append('full_name')

    if 'language' in data:
        if data['language'] not in _ALLOWED_LANGS:
            return jsonify({'error': 'Idioma no válido'}), 400
        current_user.language = data['language']
        changed.append('language')

    if 'theme' in data:
        if data['theme'] not in _ALLOWED_THEMES:
            return jsonify({'error': 'Tema no válido'}), 400
        current_user.theme = data['theme']
        changed.append('theme')

    if 'avatar_color' in data:
        color = (data['avatar_color'] or '').strip()
        if not _COLOR_RE.match(color):
            return jsonify({'error': 'Color no válido (formato #RRGGBB)'}), 400
        current_user.avatar_color = color
        changed.append('avatar_color')

    if 'email_notifications' in data:
        current_user.email_notifications = bool(data['email_notifications'])
        changed.append('email_notifications')

    if not changed:
        return jsonify({'error': 'No se enviaron campos para actualizar'}), 400

    current_user.updated_at = datetime.utcnow()
    db.session.commit()
    log_action('profile_updated', f'user:{current_user.username}', f'fields={changed}')
    return jsonify(current_user.to_dict())


# Eventos configurables por usuario (deben coincidir con las columnas de
# NotificationPreference y los defaults de modules/notifications/dispatch.py).
_NOTIF_EVENTS_DEFAULTS = {
    'machine_offline':  True,
    'service_down':     True,
    'backup_failed':    True,
    'backup_completed': True,
    'daily_summary':    False,
}


def _notif_prefs_dict(pref):
    """Devuelve un dict con las preferencias del usuario, o los defaults
    del modelo si aún no tiene fila en NotificationPreference."""
    if pref is None:
        return dict(_NOTIF_EVENTS_DEFAULTS)
    return {k: bool(getattr(pref, k, default))
            for k, default in _NOTIF_EVENTS_DEFAULTS.items()}


@auth_bp.route('/api/profile/notifications', methods=['GET'])
@login_required
def get_notification_prefs():
    pref = NotificationPreference.query.filter_by(user_id=current_user.id).first()
    return jsonify(_notif_prefs_dict(pref))


@auth_bp.route('/api/profile/notifications', methods=['POST'])
@login_required
def update_notification_prefs():
    data = request.get_json(silent=True) or {}
    pref = NotificationPreference.query.filter_by(user_id=current_user.id).first()
    if pref is None:
        pref = NotificationPreference(user_id=current_user.id)
        db.session.add(pref)

    changed = []
    for field in _NOTIF_EVENTS_DEFAULTS:
        if field in data:
            new_val = bool(data[field])
            if getattr(pref, field) != new_val:
                setattr(pref, field, new_val)
                changed.append(field)

    if not changed:
        return jsonify(_notif_prefs_dict(pref))

    pref.updated_at = datetime.utcnow()
    db.session.commit()
    log_action(
        'notification_prefs_updated',
        f'user:{current_user.username}',
        f'fields={changed}',
    )
    return jsonify(_notif_prefs_dict(pref))


@auth_bp.route('/api/profile/password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json(silent=True) or {}
    current_pw = data.get('current_password', '')
    new_pw = data.get('new_password', '')
    confirm_pw = data.get('confirm_password', '')

    if not current_user.check_password(current_pw):
        return jsonify({'error': 'La contraseña actual es incorrecta'}), 400
    if new_pw != confirm_pw:
        return jsonify({'error': 'Las contraseñas nuevas no coinciden'}), 400
    try:
        validate_password_strength(new_pw)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    current_user.set_password(new_pw)
    current_user.updated_at = datetime.utcnow()
    db.session.commit()
    log_action('password_changed', f'user:{current_user.username}')

    # Aviso al usuario por email (no respeta preferencias, motivo: seguridad).
    # Import diferido: ver justificación en create_user() (ciclo con
    # modules.notifications). Mantener al final del handler para no afectar
    # al cambio de contraseña si el módulo de notificaciones falla.
    try:
        from modules.notifications.dispatch import notify_password_changed
        notify_password_changed(current_user, ip=request.remote_addr)
    except Exception:                              # noqa: BLE001
        pass

    return jsonify({'message': 'Contraseña cambiada correctamente'})
