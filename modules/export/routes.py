import io
import json
from datetime import datetime, timedelta

from flask import Blueprint, request, send_file, jsonify
from flask_login import login_required

from config import Config
from utils.powershell_client import resolve_and_send

export_bp = Blueprint('export', __name__)

# ── helpers ───────────────────────────────────────────────────────────────────


def _mid():
    return request.args.get('machine_id') or None


def _send(cmd):
    return resolve_and_send(cmd, _mid())


def _parse_kv(text):
    text = (text or '').lstrip('﻿')
    data = {}
    for line in text.strip().splitlines():
        if ': ' in line:
            k, v = line.split(': ', 1)
            data[k.strip()] = v.strip()
    return data


def _parse_pipe(text):
    text = (text or '').lstrip('﻿')
    lines = text.strip().splitlines()
    if not lines:
        return []
    headers = [h.strip() for h in lines[0].split('|')]
    rows = []
    for line in lines[1:]:
        line = line.strip()
        if not line or '|' not in line:
            continue
        parts = line.split('|')
        rows.append({h: (parts[i].strip() if i < len(parts) else '') for i, h in enumerate(headers)})
    return rows


def _parse_blocks(text):
    """Parse multi-block text (blocks separated by blank lines) into list of dicts."""
    text = (text or '').lstrip('﻿')
    rows = []
    for block in text.strip().split('\n\n'):
        lines = block.strip().splitlines()
        if not lines:
            continue
        d = {}
        for ln in lines:
            if ': ' in ln:
                k, v = ln.split(': ', 1)
                d[k.strip()] = v.strip()
        if d:
            rows.append(d)
    return rows


def _get_machine(machine_id):
    from database.models import Machine
    if machine_id:
        try:
            return Machine.query.get(int(machine_id))
        except (ValueError, TypeError):
            return None
    return Machine.query.filter_by(
        ip=Config.POWERSHELL_SERVER,
        port=Config.POWERSHELL_PORT,
    ).first()


def _response(buf, fmt, filename):
    mime = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        if fmt == 'xlsx' else 'application/pdf'
    )
    return send_file(buf, mimetype=mime, as_attachment=True,
                     download_name=f'{filename}.{fmt}')


# ── Excel builder ─────────────────────────────────────────────────────────────

def _make_xlsx(sheets):
    """sheets: list of (title, headers, rows) — rows are lists of values."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BLUE = '0A67DC'
    LGRAY = 'F4F7FB'
    GRAY = '374151'

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor=BLUE)
    hdr_align = Alignment(horizontal='center', vertical='center')
    row_font = Font(color=GRAY, size=10)
    alt_fill = PatternFill('solid', fgColor=LGRAY)
    thin = Side(style='thin', color='E5E7EB')
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_title, headers, rows in sheets:
        ws = wb.create_sheet(title=sheet_title[:31])
        ws.row_dimensions[1].height = 22

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, brd

        for ri, row in enumerate(rows, 2):
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=str(val) if val is not None else '')
                cell.font = row_font
                cell.border = brd
                cell.alignment = Alignment(vertical='center')
                if fill:
                    cell.fill = fill

        for ci in range(1, len(headers) + 1):
            max_len = max(
                (len(str(ws.cell(r, ci).value or '')) for r in range(1, len(rows) + 2)),
                default=10,
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 55)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF builder ───────────────────────────────────────────────────────────────

def _make_pdf(title, machine_name, sections):
    """sections: list of (heading, headers, rows)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )

    BLUE = colors.HexColor('#0A67DC')
    LGRAY = colors.HexColor('#F4F7FB')
    DGRAY = colors.HexColor('#374151')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Normal'], fontSize=18,
                        textColor=BLUE, spaceAfter=4, fontName='Helvetica-Bold')
    h2 = ParagraphStyle('h2', parent=styles['Normal'], fontSize=12,
                        textColor=DGRAY, spaceAfter=6, spaceBefore=10, fontName='Helvetica-Bold')
    meta = ParagraphStyle('meta', parent=styles['Normal'], fontSize=9,
                          textColor=colors.gray, spaceAfter=10)

    page_w = A4[0] - 4 * cm
    story = [
        Paragraph(title, h1),
        Paragraph(
            f'Máquina: {machine_name} &nbsp;&#124;&nbsp; '
            f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', meta
        ),
        HRFlowable(width='100%', thickness=1, color=BLUE, spaceAfter=10),
    ]

    # Estilo de celda con wrap automático (rompe palabras largas como MTFDKBA1T0QF...)
    cell_style = ParagraphStyle(
        'cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, leading=10,
        textColor=DGRAY, wordWrap='CJK',
    )
    head_style = ParagraphStyle(
        'cellHead', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8, leading=10,
        textColor=colors.white, alignment=1, wordWrap='CJK',
    )

    def _cell(v, style=cell_style):
        s = '' if v is None else str(v)
        s = s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        return Paragraph(s, style)

    for heading, headers, rows in sections:
        story.append(Paragraph(heading, h2))
        if not rows:
            story.append(Paragraph('Sin datos disponibles.', styles['Normal']))
            story.append(Spacer(1, 0.3 * cm))
            continue
        col_w = page_w / len(headers)
        data = [[_cell(h, head_style) for h in headers]] + \
               [[_cell(v) for v in r] for r in rows]
        t = Table(data, colWidths=[col_w] * len(headers), repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), BLUE),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, 0), 5),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LGRAY]),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
            ('TOPPADDING',    (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4 * cm))

    doc.build(story)
    buf.seek(0)
    return buf


# ── /api/export/hardware ──────────────────────────────────────────────────────

@export_bp.route('/hardware')
@login_required
def export_hardware():
    fmt = request.args.get('format', 'xlsx').lower()
    if fmt not in ('xlsx', 'pdf'):
        return jsonify({'error': 'Formato no válido. Use xlsx o pdf'}), 400

    # Soporta múltiples machine_id (?machine_id=1&machine_id=2). Si no se da
    # ninguno, usa la máquina por defecto (resuelve a local).
    mids = request.args.getlist('machine_id')
    machines = []
    if mids:
        for mid in mids:
            m = _get_machine(mid)
            if m and m not in machines:
                machines.append(m)
    if not machines:
        m = _get_machine(None)
        if m:
            machines.append(m)
    if not machines:
        return jsonify({'error': 'No hay máquinas para exportar'}), 400

    ts = datetime.now().strftime('%Y%m%d_%H%M')
    fname = (f'hardware_multi_{len(machines)}_{ts}'
             if len(machines) > 1 else f'hardware_{machines[0].name}_{ts}')

    # Cargar datos por máquina (BD primero, en vivo como fallback)
    from database.models import HardwareSnapshot

    def _load(machine):
        snap = HardwareSnapshot.query.filter_by(machine_id=machine.id).first()
        if snap:
            return {
                'name':  machine.name,
                'sys':   json.loads(snap.system_json or '{}'),
                'cpu':   json.loads(snap.cpu_json or '{}'),
                'mem':   json.loads(snap.memory_json or '{}'),
                'disks': json.loads(snap.disks_json or '[]'),
                'gpu':   json.loads(snap.gpu_json or '[]'),
                'mb':    json.loads(snap.motherboard_json or '{}'),
            }
        mid = str(machine.id)
        return {
            'name':  machine.name,
            'sys':   _parse_kv(resolve_and_send('hardware_system',       mid)),
            'cpu':   _parse_kv(resolve_and_send('hardware_cpu',          mid)),
            'mem':   _parse_kv(resolve_and_send('hardware_memory',       mid)),
            'disks': _parse_pipe(resolve_and_send('hardware_disks',      mid)),
            'gpu':   _parse_pipe(resolve_and_send('hardware_gpu',        mid)),
            'mb':    _parse_kv(resolve_and_send('hardware_motherboard',  mid)),
        }

    all_data = [_load(m) for m in machines]
    multi = len(all_data) > 1

    def kv_rows(d): return [[k, v] for k, v in d.items()]

    def pipe_rows(lst): return (
        (list(lst[0].keys()), [[r.get(c, '') for c in lst[0].keys()] for r in lst])
        if lst and isinstance(lst[0], dict) else (['Campo'], [['Sin datos']])
    )

    def _safe_sheet_name(name):
        # Excel limita a 31 caracteres y prohíbe : \ / ? * [ ]
        for ch in r':\/?*[]':
            name = name.replace(ch, '_')
        return name[:31]

    if fmt == 'xlsx':
        sheets = []
        for d in all_data:
            prefix = f"{d['name']} - " if multi else ""
            dh, dr = pipe_rows(d['disks'])
            gh, gr = pipe_rows(d['gpu'])
            sheets.extend([
                (_safe_sheet_name(f"{prefix}Sistema"),    ['Campo', 'Valor'], kv_rows(d['sys']) or [['—', '—']]),
                (_safe_sheet_name(f"{prefix}CPU"),        ['Campo', 'Valor'], kv_rows(d['cpu']) or [['—', '—']]),
                (_safe_sheet_name(f"{prefix}Memoria"),    ['Campo', 'Valor'], kv_rows(d['mem']) or [['—', '—']]),
                (_safe_sheet_name(f"{prefix}Discos"),     dh, dr),
                (_safe_sheet_name(f"{prefix}GPU"),        gh, gr),
                (_safe_sheet_name(f"{prefix}Placa base"), ['Campo', 'Valor'], kv_rows(d['mb']) or [['—', '—']]),
            ])
        return _response(_make_xlsx(sheets), 'xlsx', fname)

    sections = []
    for d in all_data:
        prefix = f"[{d['name']}] " if multi else ""
        dh, dr = pipe_rows(d['disks'])
        gh, gr = pipe_rows(d['gpu'])
        sections.extend([
            (f"{prefix}Sistema",    ['Campo', 'Valor'], kv_rows(d['sys']) or [['—', '—']]),
            (f"{prefix}CPU",        ['Campo', 'Valor'], kv_rows(d['cpu']) or [['—', '—']]),
            (f"{prefix}Memoria",    ['Campo', 'Valor'], kv_rows(d['mem']) or [['—', '—']]),
            (f"{prefix}Discos",     dh, dr),
            (f"{prefix}GPU",        gh, gr),
            (f"{prefix}Placa base", ['Campo', 'Valor'], kv_rows(d['mb']) or [['—', '—']]),
        ])
    subtitle = ', '.join(d['name'] for d in all_data) if len(all_data) <= 3 else f"{len(all_data)} máquinas"
    return _response(_make_pdf('Informe de Hardware', subtitle, sections), 'pdf', fname)


# ── /api/export/rendimiento ───────────────────────────────────────────────────

@export_bp.route('/rendimiento')
@login_required
def export_rendimiento():
    fmt = request.args.get('format', 'xlsx').lower()
    days = min(int(request.args.get('days', 7)), 30)
    if fmt not in ('xlsx', 'pdf'):
        return jsonify({'error': 'Formato no válido. Use xlsx o pdf'}), 400

    machine = _get_machine(_mid())
    mname = machine.name if machine else 'Local'
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    fname = f'rendimiento_{mname}_{ts}'

    from database.models import SystemMetric
    since = datetime.utcnow() - timedelta(days=days)
    metrics = []
    if machine:
        metrics = (
            SystemMetric.query
            .filter(SystemMetric.machine_id == machine.id,
                    SystemMetric.timestamp >= since)
            .order_by(SystemMetric.timestamp)
            .all()
        )

    headers = ['Fecha/Hora', 'CPU (%)', 'Memoria (%)', 'Disco usado (GB)', 'Disco libre (GB)']
    rows = [
        [
            m.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
            round(m.cpu or 0, 2),
            round(m.memory or 0, 2),
            round(m.disk_used or 0, 2),
            round(m.disk_free or 0, 2),
        ]
        for m in metrics
    ] or [['Sin datos en el período seleccionado', '', '', '', '']]

    if fmt == 'xlsx':
        return _response(
            _make_xlsx([(f'Métricas ({days}d)', headers, rows)]),
            'xlsx', fname,
        )
    return _response(
        _make_pdf(f'Informe de Rendimiento — últimos {days} días', mname,
                  [(f'Historial ({days} días)', headers, rows)]),
        'pdf', fname,
    )


# ── /api/export/particiones ───────────────────────────────────────────────────

@export_bp.route('/particiones')
@login_required
def export_particiones():
    fmt = request.args.get('format', 'xlsx').lower()
    if fmt not in ('xlsx', 'pdf'):
        return jsonify({'error': 'Formato no válido. Use xlsx o pdf'}), 400

    machine = _get_machine(_mid())
    mname = machine.name if machine else 'Local'
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    fname = f'particiones_{mname}_{ts}'

    disks = _parse_blocks(_send('list_disks'))
    parts = _parse_blocks(_send('list_partitions'))

    dh = ['Disco', 'Modelo', 'Tamaño', 'Estado', 'Interfaz', 'Nº Serie']
    dr = [[d.get('Disco', ''), d.get('Modelo', ''), d.get('Tamaño', ''),
           d.get('Estado', ''), d.get('Interfaz', ''), d.get('Número de Serie', '')] for d in disks] \
        or [['Sin datos'] * len(dh)]

    ph = ['Partición', 'Letra', 'Etiqueta', 'Sistema arch.', 'Tamaño', 'Usado', 'Tipo']
    pr = [[p.get('Partición', ''), p.get('Letra', ''), p.get('Etiqueta', ''),
           p.get('Sistema de archivos', ''), p.get('Tamaño', ''),
           p.get('Usado', ''), p.get('Tipo', '')] for p in parts] \
        or [['Sin datos'] * len(ph)]

    if fmt == 'xlsx':
        return _response(_make_xlsx([('Discos', dh, dr), ('Particiones', ph, pr)]), 'xlsx', fname)
    return _response(
        _make_pdf('Informe de Particiones', mname,
                  [('Discos físicos', dh, dr), ('Particiones', ph, pr)]),
        'pdf', fname,
    )


# ── /api/export/red ───────────────────────────────────────────────────────────

@export_bp.route('/red')
@login_required
def export_red():
    fmt = request.args.get('format', 'xlsx').lower()
    if fmt not in ('xlsx', 'pdf'):
        return jsonify({'error': 'Formato no válido. Use xlsx o pdf'}), 400

    machine = _get_machine(_mid())
    mname = machine.name if machine else 'Local'
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    fname = f'red_{mname}_{ts}'

    adapters = _parse_pipe(_send('net_adapters'))
    stats = _parse_pipe(_send('net_stats'))

    ah = ['Nombre', 'Tipo', 'Estado', 'IP', 'Velocidad']
    ar = [[a.get('name', ''), a.get('type', ''), a.get('status', ''),
           a.get('ip', ''), a.get('speed', '')] for a in adapters] \
        or [['Sin datos'] * len(ah)]

    sh = ['Adaptador', 'Recibido (MB)', 'Enviado (MB)', 'Pkt. RX', 'Pkt. TX']
    sr = [[s.get('name', ''), s.get('rx_mb', ''), s.get('tx_mb', ''),
           s.get('rx_packets', ''), s.get('tx_packets', '')] for s in stats] \
        or [['Sin datos'] * len(sh)]

    if fmt == 'xlsx':
        return _response(_make_xlsx([('Adaptadores', ah, ar), ('Estadísticas', sh, sr)]), 'xlsx', fname)
    return _response(
        _make_pdf('Informe de Red', mname,
                  [('Adaptadores de red', ah, ar), ('Estadísticas', sh, sr)]),
        'pdf', fname,
    )
